import sqlite3
import os
from datetime import datetime


class DatabaseManager:
    def __init__(self, db_path=None):
        if db_path is None:
         db_path = os.path.join("data", "notas.db")
        self.db_path = db_path
        os.makedirs(os.path.dirname(db_path), exist_ok=True)
        self.init_database()
    
    def get_connection(self):
        return sqlite3.connect(self.db_path)
    
    def init_database(self):
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # Tabla de cursos
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS cursos (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                nombre TEXT UNIQUE NOT NULL,
                descripcion TEXT,
                total_estudiantes INTEGER DEFAULT 0,
                fecha_creacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Tabla de evaluaciones
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS evaluaciones (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                curso_id INTEGER NOT NULL,
                nombre TEXT NOT NULL,
                porcentaje REAL NOT NULL,
                orden INTEGER NOT NULL,
                fecha_evaluacion DATE,
                FOREIGN KEY (curso_id) REFERENCES cursos(id),
                UNIQUE(curso_id, nombre)
            )
        ''')
        
        # Tabla de estudiantes
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS estudiantes (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                curso_id INTEGER NOT NULL,
                nombre TEXT NOT NULL,
                grupo INTEGER DEFAULT 1,
                email TEXT,
                FOREIGN KEY (curso_id) REFERENCES cursos(id)
            )
        ''')
        
        # Tabla de notas
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS notas (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                estudiante_id INTEGER NOT NULL,
                evaluacion_id INTEGER NOT NULL,
                nota REAL,
                fecha_registro TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                observaciones TEXT,
                FOREIGN KEY (estudiante_id) REFERENCES estudiantes(id),
                FOREIGN KEY (evaluacion_id) REFERENCES evaluaciones(id),
                UNIQUE(estudiante_id, evaluacion_id)
            )
        ''')
        
        # Tabla de configuración de sincronización
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS sync_config (
                id INTEGER PRIMARY KEY CHECK (id = 1),
                last_sync TIMESTAMP,
                drive_folder_id TEXT,
                auto_sync BOOLEAN DEFAULT 0
            )
        ''')
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS clases (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                curso_id INTEGER NOT NULL,
                encabezado TEXT NOT NULL,
                topicos TEXT,
                contenido TEXT,
                observaciones TEXT,
                fecha_clase DATE,
                fecha_modificacion TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (curso_id) REFERENCES cursos(id)
            )
        ''')
        
        # Tabla de enlaces de clases
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS clase_links (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                clase_id INTEGER NOT NULL,
                nombre TEXT,
                url TEXT,
                FOREIGN KEY (clase_id) REFERENCES clases(id) ON DELETE CASCADE
            )
        ''')
        conn.commit()
        conn.close()
    
    # ========== GESTIÓN DE CURSOS ==========
    
    def crear_curso(self, nombre, descripcion=""):
        """Crea un nuevo curso"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('''
                INSERT INTO cursos (nombre, descripcion)
                VALUES (?, ?)
            ''', (nombre, descripcion))
            conn.commit()
            return cursor.lastrowid, None
        except sqlite3.IntegrityError:
            return None, "Ya existe un curso con ese nombre"
        finally:
            conn.close()
    
    def actualizar_curso(self, curso_id, nombre=None, descripcion=None):
        """Actualiza datos de un curso"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            if nombre:
                cursor.execute('UPDATE cursos SET nombre = ? WHERE id = ?', (nombre, curso_id))
            if descripcion is not None:
                cursor.execute('UPDATE cursos SET descripcion = ? WHERE id = ?', (descripcion, curso_id))
            conn.commit()
            return True, None
        except sqlite3.IntegrityError:
            return False, "Ya existe un curso con ese nombre"
        except Exception as e:
            return False, str(e)
        finally:
            conn.close()
    
    def eliminar_curso(self, curso_id):
        """Elimina un curso y todo su contenido"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            # Eliminar notas relacionadas
            cursor.execute('''
                DELETE FROM notas WHERE evaluacion_id IN 
                (SELECT id FROM evaluaciones WHERE curso_id = ?)
            ''', (curso_id,))
            
            # Eliminar evaluaciones
            cursor.execute('DELETE FROM evaluaciones WHERE curso_id = ?', (curso_id,))
            
            # Eliminar estudiantes
            cursor.execute('DELETE FROM estudiantes WHERE curso_id = ?', (curso_id,))
            
            # Eliminar curso
            cursor.execute('DELETE FROM cursos WHERE id = ?', (curso_id,))
            
            conn.commit()
            return True, None
        except Exception as e:
            return False, str(e)
        finally:
            conn.close()
    
    def get_cursos(self):
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT c.id, c.nombre, c.descripcion, 
                   COUNT(DISTINCT e.id) as total_estudiantes,
                   COUNT(DISTINCT ev.id) as total_evaluaciones
            FROM cursos c
            LEFT JOIN estudiantes e ON c.id = e.curso_id
            LEFT JOIN evaluaciones ev ON c.id = ev.curso_id
            GROUP BY c.id
            ORDER BY c.fecha_creacion DESC
        ''')
        cursos = cursor.fetchall()
        conn.close()
        return cursos
    
    # ========== GESTIÓN DE EVALUACIONES ==========
    
    def agregar_evaluacion(self, curso_id, nombre, porcentaje, orden=None):
        """Agrega una evaluación a un curso"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # Si no se especifica orden, poner al final
        if orden is None:
            cursor.execute('''
                SELECT COALESCE(MAX(orden), 0) + 1 FROM evaluaciones WHERE curso_id = ?
            ''', (curso_id,))
            orden = cursor.fetchone()[0]
        
        try:
            cursor.execute('''
                INSERT INTO evaluaciones (curso_id, nombre, porcentaje, orden)
                VALUES (?, ?, ?, ?)
            ''', (curso_id, nombre, porcentaje, orden))
            conn.commit()
            return cursor.lastrowid, None
        except sqlite3.IntegrityError:
            return None, "Ya existe una evaluación con ese nombre en este curso"
        finally:
            conn.close()
    
    def actualizar_evaluacion(self, evaluacion_id, nombre=None, porcentaje=None):
        """Actualiza datos de una evaluación"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            if nombre:
                cursor.execute('UPDATE evaluaciones SET nombre = ? WHERE id = ?', (nombre, evaluacion_id))
            if porcentaje is not None:
                cursor.execute('UPDATE evaluaciones SET porcentaje = ? WHERE id = ?', (porcentaje, evaluacion_id))
            conn.commit()
            return True, None
        except sqlite3.IntegrityError:
            return False, "Ya existe una evaluación con ese nombre"
        except Exception as e:
            return False, str(e)
        finally:
            conn.close()
    
    def eliminar_evaluacion(self, evaluacion_id):
        """Elimina una evaluación y sus notas asociadas"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('DELETE FROM notas WHERE evaluacion_id = ?', (evaluacion_id,))
            cursor.execute('DELETE FROM evaluaciones WHERE id = ?', (evaluacion_id,))
            conn.commit()
            return True, None
        except Exception as e:
            return False, str(e)
        finally:
            conn.close()
    
    def actualizar_orden_evaluaciones(self, curso_id, orden_ids):
        """Actualiza el orden de las evaluaciones"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            for nuevo_orden, eval_id in enumerate(orden_ids, 1):
                cursor.execute('''
                    UPDATE evaluaciones SET orden = ? WHERE id = ? AND curso_id = ?
                ''', (nuevo_orden, eval_id, curso_id))
            conn.commit()
            return True
        except Exception as e:
            return False
        finally:
            conn.close()
    
    def get_evaluaciones(self, curso_id):
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, nombre, porcentaje, orden, fecha_evaluacion 
            FROM evaluaciones 
            WHERE curso_id = ? 
            ORDER BY orden
        ''', (curso_id,))
        evals = cursor.fetchall()
        conn.close()
        return evals

    def verificar_porcentaje_total(self, curso_id):
        """Verifica que el total de puntos sea 100"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT COALESCE(SUM(porcentaje), 0) FROM evaluaciones WHERE curso_id = ?
        ''', (curso_id,))
        total = cursor.fetchone()[0]
        conn.close()
        return total
    
    # ========== GESTIÓN DE ESTUDIANTES ==========
    
    def agregar_estudiante(self, curso_id, nombre, grupo=1, email=None):
        """Agrega un estudiante a un curso"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('''
                INSERT INTO estudiantes (curso_id, nombre, grupo, email)
                VALUES (?, ?, ?, ?)
            ''', (curso_id, nombre, grupo, email))
            conn.commit()
            return cursor.lastrowid, None
        except Exception as e:
            return None, str(e)
        finally:
            conn.close()
    
    def actualizar_estudiante(self, estudiante_id, nombre=None, grupo=None, email=None):
        """Actualiza datos de un estudiante"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            if nombre:
                cursor.execute('UPDATE estudiantes SET nombre = ? WHERE id = ?', (nombre, estudiante_id))
            if grupo is not None:
                cursor.execute('UPDATE estudiantes SET grupo = ? WHERE id = ?', (grupo, estudiante_id))
            if email is not None:
                cursor.execute('UPDATE estudiantes SET email = ? WHERE id = ?', (email, estudiante_id))
            conn.commit()
            return True, None
        except Exception as e:
            return False, str(e)
        finally:
            conn.close()
    
    def eliminar_estudiante(self, estudiante_id):
        """Elimina un estudiante y sus notas"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('DELETE FROM notas WHERE estudiante_id = ?', (estudiante_id,))
            cursor.execute('DELETE FROM estudiantes WHERE id = ?', (estudiante_id,))
            conn.commit()
            return True, None
        except Exception as e:
            return False, str(e)
        finally:
            conn.close()
    
    def get_estudiantes(self, curso_id, grupo=None):
        conn = self.get_connection()
        cursor = conn.cursor()
        if grupo:
            cursor.execute('''
                SELECT id, nombre, grupo, email 
                FROM estudiantes 
                WHERE curso_id = ? AND grupo = ?
                ORDER BY nombre
            ''', (curso_id, grupo))
        else:
            cursor.execute('''
                SELECT id, nombre, grupo, email 
                FROM estudiantes 
                WHERE curso_id = ?
                ORDER BY grupo, nombre
            ''', (curso_id,))
        ests = cursor.fetchall()
        conn.close()
        return ests
    
    # ========== GESTIÓN DE NOTAS ==========
    
    def get_nota(self, estudiante_id, evaluacion_id):
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT nota, observaciones 
            FROM notas 
            WHERE estudiante_id = ? AND evaluacion_id = ?
        ''', (estudiante_id, evaluacion_id))
        result = cursor.fetchone()
        conn.close()
        return result if result else (None, "")
    
    def guardar_nota(self, estudiante_id, evaluacion_id, nota, observaciones=""):
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            INSERT INTO notas (estudiante_id, evaluacion_id, nota, observaciones)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(estudiante_id, evaluacion_id) 
            DO UPDATE SET nota=excluded.nota, observaciones=excluded.observaciones, 
                         fecha_registro=CURRENT_TIMESTAMP
        ''', (estudiante_id, evaluacion_id, nota, observaciones))
        conn.commit()
        conn.close()

    def calcular_promedio(self, estudiante_id, curso_id):
        """Calcula el promedio sumando directamente los puntos obtenidos"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # Obtener todas las evaluaciones del curso con sus puntos máximos
        cursor.execute('''
            SELECT id, porcentaje 
            FROM evaluaciones 
            WHERE curso_id = ?
        ''', (curso_id,))
        evaluaciones = {row[0]: row[1] for row in cursor.fetchall()}
        
        if not evaluaciones:
            conn.close()
            return 0.0, 0
        
        # Obtener las notas del estudiante
        cursor.execute('''
            SELECT evaluacion_id, nota 
            FROM notas 
            WHERE estudiante_id = ? AND nota IS NOT NULL
        ''', (estudiante_id,))
        notas_registradas = {row[0]: row[1] for row in cursor.fetchall()}
        
        conn.close()
        
        # Sumar puntos obtenidos directamente
        puntos_obtenidos = 0
        puntos_maximos_posibles = sum(evaluaciones.values())  # Debería ser 100
        
        for eval_id, puntos_max in evaluaciones.items():
            if eval_id in notas_registradas:
                puntos_obtenidos += notas_registradas[eval_id]
        
        return round(puntos_obtenidos, 2), int(puntos_maximos_posibles)
    
        # ========== GESTIÓN DE CLASES ==========
    
    def crear_clase(self, curso_id, encabezado, topicos="", contenido="", observaciones="", fecha_clase=None):
        """Crea una nueva clase en la base de datos"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('''
                INSERT INTO clases (curso_id, encabezado, topicos, contenido, observaciones, fecha_clase)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (curso_id, encabezado, topicos, contenido, observaciones, fecha_clase))
            conn.commit()
            return cursor.lastrowid, None
        except Exception as e:
            return None, str(e)
        finally:
            conn.close()
    
    def actualizar_clase(self, clase_id, encabezado=None, topicos=None, contenido=None, observaciones=None, fecha_clase=None):
        """Actualiza una clase existente"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            if encabezado is not None:
                cursor.execute('UPDATE clases SET encabezado = ? WHERE id = ?', (encabezado, clase_id))
            if topicos is not None:
                cursor.execute('UPDATE clases SET topicos = ? WHERE id = ?', (topicos, clase_id))
            if contenido is not None:
                cursor.execute('UPDATE clases SET contenido = ? WHERE id = ?', (contenido, clase_id))
            if observaciones is not None:
                cursor.execute('UPDATE clases SET observaciones = ? WHERE id = ?', (observaciones, clase_id))
            if fecha_clase is not None:
                cursor.execute('UPDATE clases SET fecha_clase = ? WHERE id = ?', (fecha_clase, clase_id))
            # Actualizar fecha de modificación
            cursor.execute('UPDATE clases SET fecha_modificacion = CURRENT_TIMESTAMP WHERE id = ?', (clase_id,))
            conn.commit()
            return True, None
        except Exception as e:
            return False, str(e)
        finally:
            conn.close()
    
    def eliminar_clase(self, clase_id):
        """Elimina una clase y sus enlaces"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            # Los enlaces se eliminan automáticamente por ON DELETE CASCADE
            cursor.execute('DELETE FROM clases WHERE id = ?', (clase_id,))
            conn.commit()
            return True, None
        except Exception as e:
            return False, str(e)
        finally:
            conn.close()
    
    def get_clases(self, curso_id):
        """Obtiene todas las clases de un curso"""
        conn = self.get_connection()
        cursor = conn.cursor()
        cursor.execute('''
            SELECT id, encabezado, topicos, contenido, observaciones, fecha_clase, fecha_modificacion
            FROM clases 
            WHERE curso_id = ? 
            ORDER BY fecha_modificacion DESC
        ''', (curso_id,))
        clases = cursor.fetchall()
        conn.close()
        return clases
    
    def get_clase_por_id(self, clase_id):
        """Obtiene una clase específica con sus enlaces"""
        conn = self.get_connection()
        cursor = conn.cursor()
        
        # Obtener datos de la clase
        cursor.execute('''
            SELECT id, curso_id, encabezado, topicos, contenido, observaciones, fecha_clase, fecha_modificacion
            FROM clases WHERE id = ?
        ''', (clase_id,))
        clase = cursor.fetchone()
        
        if not clase:
            conn.close()
            return None
        
        # Obtener enlaces
        cursor.execute('SELECT nombre, url FROM clase_links WHERE clase_id = ?', (clase_id,))
        links = [{"nombre": row[0], "url": row[1]} for row in cursor.fetchall()]
        
        conn.close()
        
        return {
            "id": clase[0],
            "curso_id": clase[1],
            "encabezado": clase[2],
            "topicos": clase[3],
            "contenido": clase[4],
            "observaciones": clase[5],
            "fecha_clase": clase[6],
            "fecha_modificacion": clase[7],
            "links": links
        }
    
    def agregar_link_clase(self, clase_id, nombre, url):
        """Agrega un enlace a una clase"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('''
                INSERT INTO clase_links (clase_id, nombre, url)
                VALUES (?, ?, ?)
            ''', (clase_id, nombre, url))
            conn.commit()
            return cursor.lastrowid, None
        except Exception as e:
            return None, str(e)
        finally:
            conn.close()
    
    def eliminar_links_clase(self, clase_id):
        """Elimina todos los enlaces de una clase (útil al actualizar)"""
        conn = self.get_connection()
        cursor = conn.cursor()
        try:
            cursor.execute('DELETE FROM clase_links WHERE clase_id = ?', (clase_id,))
            conn.commit()
            return True, None
        except Exception as e:
            return False, str(e)
        finally:
            conn.close()
    
    # ========== EXPORTACIÓN ==========
    
    def exportar_a_excel(self, curso_id, filepath):
        """Exporta todas las notas del curso a Excel usando openpyxl"""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        conn = self.get_connection()
        
        cursor = conn.cursor()
        cursor.execute("SELECT nombre FROM cursos WHERE id = ?", (curso_id,))
        curso_nombre = cursor.fetchone()[0]
        
        evals = self.get_evaluaciones(curso_id)
        eval_ids = [e[0] for e in evals]
        eval_nombres = [f"{e[1]} ({e[2]}%)" for e in evals]
        
        ests = self.get_estudiantes(curso_id)
        
        # Crear libro de Excel
        wb = Workbook()
        ws = wb.active
        ws.title = 'Notas'
        
        # Encabezados
        headers = ['ID', 'Nombre', 'Grupo', 'Email'] + eval_nombres + ['PROMEDIO']
        ws.append(headers)
        
        # Formato encabezados (negrita, fondo gris)
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        header_align = Alignment(horizontal='center', vertical='center')
        
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
        
        # Datos de estudiantes
        for est in ests:
            est_id, nombre, grupo, email = est
            row = [est_id, nombre, grupo, email or '']
            
            for eval_id in eval_ids:
                nota, _ = self.get_nota(est_id, eval_id)
                row.append(nota if nota is not None else '')
            
            promedio, _ = self.calcular_promedio(est_id, curso_id)
            row.append(promedio)
            
            ws.append(row)
        
        #Prueba de código

        # Ajustar anchos de columna
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Congelar primera fila
        ws.freeze_panes = 'A2'
        
        # Guardar archivo
        wb.save(filepath)
        conn.close()
        return filepath